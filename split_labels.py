from openpyxl import load_workbook, Workbook
from collections import defaultdict

# Load the original workbook
input_file = 'your_input_file.xlsx'  # Replace with your input file name
wb = load_workbook(input_file)
sheet = wb.active

# Group data by team
team_data = defaultdict(list)

# Assuming first row is header and team column is labeled "Team"
headers = [cell.value for cell in sheet[1]]
team_column_index = headers.index('Team') + 1

for row in sheet.iter_rows(min_row=2, values_only=True):
    team = row[team_column_index - 1]
    team_data[team].append(row)

# Create a new workbook for each team and save it
for team, records in team_data.items():
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = team
    
    # Write headers
    new_sheet.append(headers)
    
    # Write team records
    for record in records:
        new_sheet.append(record)
    
    # Save the new workbook
    output_file = f'{team}_records.xlsx'
    new_wb.save(output_file)
    print(f'Created file: {output_file}')
